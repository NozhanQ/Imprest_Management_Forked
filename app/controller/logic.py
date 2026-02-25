from __future__ import annotations
from dataclasses import dataclass
from app.data.data_base import Load_Save_Data
from PyQt6.QtGui import QStandardItem, QStandardItemModel
from PyQt6.QtWidgets import (
    QWidget, QFileDialog, QMessageBox
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from typing import Iterable, Optional, Union
import ast
from pathlib import Path

from PyQt6.QtCore import Qt, QRect, QRectF
from PyQt6.QtGui import QPainter, QPen, QImageReader, QFont
from PyQt6.QtPrintSupport import QPrinter
from PyQt6.QtGui import QPageSize, QPageLayout


PathLike = Union[str, Path]



class receipt_entry_logic:
    def browse_image(
            self,
            parent: QWidget,
            title: str = "add image",
            start_dir: Optional[str | Path] = None,
    ) -> list[str]:
        start_dir_str = str(start_dir) if start_dir else ""

        file_paths, _ = QFileDialog.getOpenFileNames(
            parent,
            title,
            start_dir_str,
            "Images (*.png *.jpg *.jpeg *.bmp *.gif *.webp);;All Files (*)",
        )

        return file_paths




@dataclass
class LoginResult:
    ok: bool
    error: str = ""

class main_window_logic:
    def validate_inputs(self, username: str, password: str) -> LoginResult:
        if not username:
            return LoginResult(False, "Username is empty")
        if not password:
            return LoginResult(False, "Password is empty")
        return LoginResult(True)

    def login(self, username: str, password: str) -> LoginResult:
        # later: check DB / API
        return self.validate_inputs(username, password)



class calling_page_logic:
    def __init__(self):
        super().__init__()

    def load_invoices(self):
        # Your repo functions return (headers, rows), so we ignore headers
        if self.rbInvoiceNo.isChecked():
            rows = Load_Save_Data.get_invoices_by_Invoice_NO(self.leInvoiceNo.text().strip())

        elif self.rbTimeRange.isChecked():
            date_str_start = self.deLoginStart.date().toString("yyyy-MM-dd")
            date_str_end = self.deLoginEnd.date().toString("yyyy-MM-dd")
            rows = Load_Save_Data.get_invoices_by_time_range(date_str_start, date_str_end)

        elif self.rbRegistrationDate.isChecked():
            date_str = self.deRegstrationDate.date().toString("yyyy-MM-dd")
            rows = Load_Save_Data.get_invoices_by_regestrationdate(date_str)

        else:  # rbExplanation
            rows = Load_Save_Data.get_invoices_by_explanation(self.leExplanation.text().strip())

        # Clear + keep the static first row
        self.model.clear()
        self.model.setColumnCount(len(self.headers))
        self.model.setHorizontalHeaderLabels([""] * len(self.headers))
        self.model.appendRow([QStandardItem(h) for h in self.headers])

        # Fill DB rows cell-by-cell
        for row in rows:
            items = [QStandardItem("" if v is None else str(v)) for v in row]
            self.model.appendRow(items)


    def sending_img_path(self) -> list[str]:
        if self.rbInvoiceNo.isChecked():
            img_path = Load_Save_Data().id_by_invoice_no(self.leInvoiceNo.text().strip())

        elif self.rbTimeRange.isChecked():
            date_str_start = self.deLoginStart.date().toString("yyyy-MM-dd")
            date_str_end = self.deLoginEnd.date().toString("yyyy-MM-dd")
            img_path = Load_Save_Data().id_by_time_range(date_str_start, date_str_end)

        elif self.rbRegistrationDate.isChecked():
            date_str = self.deRegstrationDate.date().toString("yyyy-MM-dd")
            img_path = Load_Save_Data().id_by_regestrationdate(date_str)

        else:  # rbExplanation
            img_path = Load_Save_Data().id_by_explanation(self.leExplanation.text().strip())

        return img_path


    def populate_table(self, rows):
        model = QStandardItemModel()
        model.setHorizontalHeaderLabels([
            "Invoice No","amount" , "Explanation",
            "Registration Date", "Login Date"
        ])

        for row in rows:
            model.appendRow(
                [QStandardItem(str(col)) for col in row]
            )

        self.tableView.setModel(model)



    def export_tableview_to_pdf(
            self,
            table,
            filename: str,
            image_col_name: str = "image_path",
            extra_image_sources=None,
            recursive: bool = False
    ) -> None:
        model = table.model()
        if model is None:
            raise RuntimeError("Table has no model")

        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
        printer.setOutputFileName(filename)
        printer.setPageSize(QPageSize(QPageSize.PageSizeId.A4))
        printer.setPageOrientation(QPageLayout.Orientation.Portrait)
        printer.setResolution(300)

        painter = QPainter()
        if not painter.begin(printer):
            raise RuntimeError("Could not start PDF painter")

        painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
        painter.setRenderHint(QPainter.RenderHint.TextAntialiasing, True)
        painter.setPen(QPen(Qt.GlobalColor.black, 1))

        def pt_to_px_x(pt: float) -> int:
            return int(pt * printer.logicalDpiX() / 72.0)

        def pt_to_px_y(pt: float) -> int:
            return int(pt * printer.logicalDpiY() / 72.0)

        def _clean_path(s) -> str:
            return str(s).strip().strip('"').strip("'").strip()

        def _flatten_values(v):
            if v is None:
                return []
            if isinstance(v, Path):
                return [str(v)]
            if isinstance(v, str):
                s = _clean_path(v)
                if (s.startswith("(") and s.endswith(")")) or (s.startswith("[") and s.endswith("]")):
                    try:
                        parsed = ast.literal_eval(s)
                        return _flatten_values(parsed)
                    except Exception:
                        return [s]
                return [s]
            if isinstance(v, (list, tuple, set)):
                if isinstance(v, tuple) and len(v) == 1:
                    return _flatten_values(v[0])
                out = []
                for item in v:
                    out.extend(_flatten_values(item))
                return [x for x in out if x]
            return [str(v)]

        def _resolve_source_strings(raw_values, base_dir: Path):
            out = []
            for s in _flatten_values(raw_values):
                s = _clean_path(s)
                if not s:
                    continue
                p = Path(s)
                if p.is_absolute():
                    out.append(str(p))
                else:
                    if ("\\" not in s) and ("/" not in s) and (":" not in s):
                        out.append(str(base_dir / s))
                    else:
                        out.append(str(p))
            return out

        def find_col_by_header(header_name: str):
            for c in range(model.columnCount()):
                h = model.headerData(c, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole)
                if h is not None and str(h) == header_name:
                    return c
            return None

        def sources_from_model():
            col = find_col_by_header(image_col_name)
            if col is None:
                return []
            srcs = []
            for r in range(model.rowCount()):
                v = model.data(model.index(r, col), Qt.ItemDataRole.DisplayRole)
                if v:
                    srcs.append(v)
            return srcs

        def collect_images(sources, recursive_flag: bool):
            all_files = []
            for src in sources:
                p = Path(_clean_path(src))
                if p.is_dir():
                    it = p.rglob("*") if recursive_flag else p.iterdir()
                    for f in it:
                        if f.is_file():
                            all_files.append(f)
                elif p.is_file():
                    all_files.append(p)

            readable = []
            for f in sorted(all_files):
                r = QImageReader(str(f))
                r.setAutoTransform(True)
                if r.canRead():
                    readable.append(f)
            return readable

        try:
            base_family = table.font().family()
            font = QFont(base_family)
            font.setPointSize(13)
            painter.setFont(font)
            fm = painter.fontMetrics()

            page = printer.pageRect(QPrinter.Unit.DevicePixel)

            margin_x = pt_to_px_x(24)
            margin_y = pt_to_px_y(24)

            left = int(page.left() + margin_x)
            top = int(page.top() + margin_y)
            right = int(page.right() - margin_x)
            bottom = int(page.bottom() - margin_y)

            avail_w = right - left

            cell_pad = pt_to_px_x(4)
            row_h = int(fm.height() + pt_to_px_y(10))
            header_h = int(row_h + pt_to_px_y(3))

            rows = model.rowCount()
            cols = model.columnCount()

            sample_rows = min(rows, 80)
            col_widths = []
            min_w = pt_to_px_x(60)
            max_w = pt_to_px_x(200)

            for c in range(cols):
                header = str(model.headerData(c, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole) or "")
                w = fm.horizontalAdvance(header) + 2 * cell_pad

                for r in range(sample_rows):
                    val = model.data(model.index(r, c), Qt.ItemDataRole.DisplayRole)
                    txt = "" if val is None else str(val)
                    w = max(w, fm.horizontalAdvance(txt) + 2 * cell_pad)

                w = max(min_w, min(int(w), max_w))
                col_widths.append(w)

            total_w = sum(col_widths)
            if total_w > avail_w:
                scale = avail_w / total_w
                col_widths = [max(pt_to_px_x(40), int(w * scale)) for w in col_widths]
            elif total_w < avail_w and cols > 0:
                extra = (avail_w - total_w) // cols
                col_widths = [w + extra for w in col_widths]

            def draw_table_header(y: int) -> int:
                x = left
                for c in range(cols):
                    w = col_widths[c]
                    rect = QRect(int(x), int(y), int(w), int(header_h))
                    painter.drawRect(rect)
                    text = str(model.headerData(c, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole) or "")
                    painter.drawText(
                        rect.adjusted(cell_pad, 0, -cell_pad, 0),
                        Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft,
                        text
                    )
                    x += w
                return y + header_h

            y = top
            y = draw_table_header(y)

            for r in range(rows):
                if y + row_h > bottom:
                    printer.newPage()
                    page2 = printer.pageRect(QPrinter.Unit.DevicePixel)

                    left = int(page2.left() + margin_x)
                    top = int(page2.top() + margin_y)
                    right = int(page2.right() - margin_x)
                    bottom = int(page2.bottom() - margin_y)

                    y = top
                    y = draw_table_header(y)

                x = left
                for c in range(cols):
                    w = col_widths[c]
                    rect = QRect(int(x), int(y), int(w), int(row_h))
                    painter.drawRect(rect)

                    val = model.data(model.index(r, c), Qt.ItemDataRole.DisplayRole)
                    text = "" if val is None else str(val)

                    align = Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft
                    try:
                        float(text)
                        align = Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignRight
                    except Exception:
                        pass

                    painter.drawText(rect.adjusted(cell_pad, 0, -cell_pad, 0), align, text)
                    x += w
                y += row_h

            base_dir = Path(r"D:\Work\Imprest_Management\Imprest_Management_Forked\image_records")

            sources = []
            sources.extend(_resolve_source_strings(sources_from_model(), base_dir))

            if extra_image_sources:
                sources.extend(_resolve_source_strings(extra_image_sources, base_dir))

            image_files = collect_images(sources, recursive)

            if not image_files:
                id_ = self.sending_img_path()
                fallback_sources = _resolve_source_strings(id_, base_dir)
                sources.extend(fallback_sources)
                image_files = collect_images(fallback_sources, recursive_flag=True)

            if image_files:
                printer.newPage()

                page_img = printer.pageRect(QPrinter.Unit.DevicePixel)
                left_i = int(page_img.left() + margin_x)
                top_i = int(page_img.top() + margin_y)
                right_i = int(page_img.right() - margin_x)
                bottom_i = int(page_img.bottom() - margin_y)
                target = QRectF(left_i, top_i, right_i - left_i, bottom_i - top_i)

                for idx, img_path in enumerate(image_files):
                    if idx > 0:
                        printer.newPage()
                        page_img = printer.pageRect(QPrinter.Unit.DevicePixel)
                        left_i = int(page_img.left() + margin_x)
                        top_i = int(page_img.top() + margin_y)
                        right_i = int(page_img.right() - margin_x)
                        bottom_i = int(page_img.bottom() - margin_y)
                        target = QRectF(left_i, top_i, right_i - left_i, bottom_i - top_i)

                    reader = QImageReader(str(img_path))
                    reader.setAutoTransform(True)
                    try:
                        reader.setDecideFormatFromContent(True)
                    except Exception:
                        pass

                    if not reader.canRead():
                        painter.drawText(
                            QRect(int(target.x()), int(target.y()), int(target.width()), pt_to_px_y(24)),
                            Qt.AlignmentFlag.AlignLeft,
                            f"Cannot read: {img_path}"
                        )
                        continue

                    img = reader.read()
                    if img.isNull():
                        painter.drawText(
                            QRect(int(target.x()), int(target.y()), int(target.width()), pt_to_px_y(24)),
                            Qt.AlignmentFlag.AlignLeft,
                            f"Null image: {img_path}"
                        )
                        continue

                    iw, ih = img.width(), img.height()
                    scale = min(target.width() / iw, target.height() / ih)
                    dw, dh = iw * scale, ih * scale
                    x = target.x() + (target.width() - dw) / 2
                    y = target.y() + (target.height() - dh) / 2
                    dest = QRectF(x, y, dw, dh)
                    painter.drawImage(dest, img)

            else:
                printer.newPage()
                page_dbg = printer.pageRect(QPrinter.Unit.DevicePixel)
                msg_rect = QRect(
                    int(page_dbg.left() + margin_x),
                    int(page_dbg.top() + margin_y),
                    int(page_dbg.width() - 2 * margin_x),
                    int(pt_to_px_y(200))
                )
                painter.drawText(
                    msg_rect,
                    Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop,
                    "No photos found!\nSources:\n- " + "\n- ".join(map(str, sources))
                )

        finally:
            painter.end()

    def export_tableview_to_excel(self, view):
        model = view.model()
        header = view.horizontalHeader()

        if model is None:
            raise RuntimeError("Could not find model")

        #columns
        export_col = []
        for visual in range(header.count()):
            logical = header.logicalIndex(visual)
            if 0<= logical < model.columnCount():
                export_col.append(logical)

        if export_col is None:
            raise RuntimeError("Could not find column")

        #building Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Export Table"

        header_font = Font(bold=True)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for out_col, logical_col in enumerate(export_col, start=1):
            title = model.headerData(logical_col, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole)
            text = "" if title is None else str(title)

            cell = ws.cell(row=1, column=out_col, value=text)
            cell.font = header_font
            cell.alignment = header_align

        ws.freeze_panes = "A2"

        #write all visible data rows
        start_excel_row = 2
        rows = model.rowCount()

        for r in range(rows):
            excel_row = start_excel_row + r

            for excel_col, logical_col in enumerate(export_col, start=1):
                idx = model.index(r, logical_col)
                value = model.data(idx, Qt.ItemDataRole.DisplayRole)
                if value is None:
                    state = model.data(idx, Qt.CheckStateRole)
                    if state is not None:
                        value = "✓" if state == Qt.Checked else "✗"
                    else:
                        value = ""

                ws.cell(row=excel_row, column=excel_col, value=str(value))

        #Fixing the width of cells for readability
        from openpyxl.utils import get_column_letter

        max_lens = [0] * len(export_col)

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=1, max_col=len(export_col)):
            for i, cell in enumerate(row):
                v = cell.value
                if v is None:
                    continue
                s = str(v)
                if len(s) > max_lens[i]:
                    max_lens[i] = len(s)

        for i, m in enumerate(max_lens, start=1):
            col_letter = get_column_letter(i)
            width = max(10, min(60, m + 2))
            ws.column_dimensions[col_letter].width = width

        # Step 6: Save dialog + save workbook + error handling
        parent = view.window()
        path, _ = QFileDialog.getSaveFileName(
            parent,
            "Save Excel",
            "Export.xlsx",
            "Excel Workbook (*.xlsx)"
        )

        if not path:
            return  # user cancelled

        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        try:
            wb.save(path)
        except PermissionError:
            QMessageBox.warning(
                self,
                "Save failed",
                "Could not write the file.\n"
                "If it's open in Excel, close it and try again, or choose another location."
            )
            return
        except OSError as e:
            QMessageBox.critical(
                self,
                "Save failed",
                f"OS error while saving the file:\n{e}"
            )
            return
        except Exception as e:
            QMessageBox.critical(
                self,
                "Save failed",
                f"Unexpected error while saving the file:\n{e}"
            )
            return
